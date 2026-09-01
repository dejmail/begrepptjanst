"""A staff member exporting selected concepts to an Excel workbook from admin."""

import io

import openpyxl
import pytest
from django.urls import reverse

from term_list.tests.factories import ConceptFactory, SynonymFactory, TaskOrdererFactory

pytestmark = pytest.mark.django_db


def export(client, concepts, attributes=None):
    ids = "&".join(str(c.id) for c in concepts)
    params = {"selected_concepts": ids}
    if attributes:
        params["attributes"] = attributes
    return client.get(reverse("export_chosen_attrs"), params)


class TestExportWorkbook:
    def test_exported_workbook_contains_the_selected_concepts_terms(
        self, admin_user_authenticated_client, concept
    ):
        """The exported .xlsx has one data row per selected concept, including its term."""
        concept.term = "Exportterm"
        concept.save()

        response = export(admin_user_authenticated_client, [concept])

        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        workbook = openpyxl.load_workbook(io.BytesIO(response.content))
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        header = rows[0]
        term_col = header.index("Term")
        exported_terms = [row[term_col] for row in rows[1:]]
        assert "Exportterm" in exported_terms

    def test_exported_response_is_a_downloadable_attachment(
        self, admin_user_authenticated_client, concept
    ):
        """The export response is served as a file attachment, not rendered inline."""
        response = export(admin_user_authenticated_client, [concept])
        assert "attachment" in response["Content-Disposition"]
        assert ".xlsx" in response["Content-Disposition"]

    def test_default_columns_are_always_included(
        self, admin_user_authenticated_client, concept
    ):
        """Id, Term, Definition and Status are always exported, even if not explicitly requested."""
        response = export(admin_user_authenticated_client, [concept])
        workbook = openpyxl.load_workbook(io.BytesIO(response.content))
        header = next(workbook.active.iter_rows(values_only=True))
        for expected in ("Id", "Term", "Definition", "Status"):
            assert expected in header

    def test_synonyms_column_lists_each_synonym_with_its_status(
        self, admin_user_authenticated_client, concept
    ):
        """A requested 'Synonyms' column formats each synonym as 'text - status', comma-separated."""
        SynonymFactory(concept=concept, synonym="Alias1", synonym_status="Tillåten")
        response = export(
            admin_user_authenticated_client, [concept], attributes=["Synonyms"]
        )
        workbook = openpyxl.load_workbook(io.BytesIO(response.content))
        rows = list(workbook.active.iter_rows(values_only=True))
        header, data_row = rows[0], rows[1]
        synonym_col = header.index("Synonyms")
        assert data_row[synonym_col] == "Alias1 - Tillåten"

    def test_task_requester_column_shows_the_requesters_name(
        self, admin_user_authenticated_client, concept
    ):
        """A requested 'Task_requester' column surfaces the name of the person who requested the term."""
        TaskOrdererFactory(concept=concept, name="Beställaren Persson")
        response = export(
            admin_user_authenticated_client, [concept], attributes=["Task_requester"]
        )
        workbook = openpyxl.load_workbook(io.BytesIO(response.content))
        rows = list(workbook.active.iter_rows(values_only=True))
        header, data_row = rows[0], rows[1]
        requester_col = header.index("Task_requester")
        assert data_row[requester_col] == "Beställaren Persson"
